"""
HSDES Ticket Creator
====================
自动从 Excel 模板读取数据，并通过 REST API 批量创建 HSDES test_case ticket。

使用方法:
    python hsdes_ticket_creator.py --excel "WO_Template.xlsx" --dry-run
    python hsdes_ticket_creator.py --excel "WO_Template.xlsx" --config hsdes_api_config.json

认证方式:
    默认从环境变量 HSDES_API_TOKEN 读取 token，并以 Authorization: Bearer <token> 发送。
    也可以在配置文件中改成其他 header，例如 X-API-Key。
"""

import argparse
import base64
import getpass
import http.cookiejar
import json
import logging
import os
import ssl
import sys
import time
import urllib.parse
import uuid
import zipfile
from pathlib import Path
from typing import Any
from urllib import error, request
from xml.etree import ElementTree as ET


DEFAULT_TOKEN_ENV = "HSDES_API_TOKEN"
DEFAULT_COOKIE_ENV = "HSDES_COOKIE"
DEFAULT_COOKIE_CACHE_ENV = "HSDES_COOKIE_CACHE_FILE"
DEFAULT_CONFIG_FILE = "hsdes_api_config.json"
DEFAULT_PAUSE_SECONDS = 1.0
DEFAULT_TIMEOUT_SECONDS = 30
DEFAULT_API_CLIENT = "HSD-ES Article"
COOKIE_CACHE_FILE = ".hsdes_cookie_cache"
DEFAULT_PRECHECK_STATUS_FILE = "hsdes_precheck_status.json"
HSDES_LOGIN_URL = "https://hsdes.intel.com/login"
HSDES_HOME_URL = "https://hsdes.intel.com/appstore/article-one/"

DEFAULT_FIELD_MAPPING = {
    "Title": "title",
    "Description": "description",
    "configuration": "configuration",
    "family": "family",
    "notify": "notify",
    "owner": "owner",
    "bios_rev": "bios_rev",
    "board_name": "board_name",
    "dimm_part_number": "dimm_part_number",
    "dimm_size": "dimm_size",
    "dimm_vendor": "dimm_vendor",
    "free_tag_1": "free_tag_1",
    "free_tag_2": "free_tag_2",
    "free_tag_3": "free_tag_3",
    "config_description": "config_description",
}

DEFAULT_STATIC_FIELDS: dict[str, Any] = {}

REQUIRED_EXCEL_COLUMNS = ["Title"]
MULTILINE_DISPLAY_FIELDS = {
    "description",
    "config_description",
    "test_case.config_description",
}
DEFERRED_UPDATE_FIELDS = {
    "test_case.free_tag_1",
    "test_case.free_tag_2",
    "test_case.free_tag_3",
}


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("hsdes_creator.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="HSDES Ticket Batch Creator (REST API)")
    parser.add_argument("--excel", required=True, help="Excel 模板文件路径")
    parser.add_argument("--sheet", default=None, help="工作表名称，默认第一张")
    parser.add_argument("--start-row", type=int, default=0, help="从第几条数据开始，默认 0")
    parser.add_argument("--config", default=DEFAULT_CONFIG_FILE, help="API 配置文件路径")
    parser.add_argument("--api-url", default=None, help="创建 ticket 的 REST API URL")
    parser.add_argument("--token", default=None, help="API token")
    parser.add_argument("--token-env", default=DEFAULT_TOKEN_ENV, help="读取 token 的环境变量名")
    parser.add_argument("--cookie", default=None, help="会话 Cookie 字符串（用于 SSO 会话认证）")
    parser.add_argument("--cookie-env", default=DEFAULT_COOKIE_ENV, help="读取 Cookie 的环境变量名")
    parser.add_argument("--cookie-cache-file", default=None, help="Cookie 缓存文件路径（默认按当前系统用户隔离）")
    parser.add_argument("--no-cookie-cache", action="store_true", help="禁用本地 Cookie 缓存读写")
    parser.add_argument("--username", default=None, help="Intel账户用户名（用于自动登录）")
    parser.add_argument("--password", default=None, help="Intel账户密码（用于自动登录）")
    parser.add_argument("--auth-header", default=None, help="认证 header 名，例如 Authorization")
    parser.add_argument("--auth-scheme", default=None, help="认证前缀，例如 Bearer")
    parser.add_argument("--subject", default=None, help="ticket subject，默认 test_case")
    parser.add_argument("--pause-seconds", type=float, default=DEFAULT_PAUSE_SECONDS, help="每次请求间隔秒数")
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT_SECONDS, help="HTTP 请求超时时间")
    parser.add_argument(
        "--precheck-status-file",
        default=None,
        help="预检查状态文件路径（可选，便于批处理判断）",
    )
    parser.add_argument("--dry-run", action="store_true", help="只构造 payload，不实际提交")
    parser.add_argument("--wizard", action="store_true", help="首次运行向导模式（交互式检查并选择执行策略）")
    parser.add_argument("--insecure", action="store_true", help="跳过 SSL 证书校验")
    parser.add_argument("--auth-debug", action="store_true", help="输出方式1认证调试日志（不打印敏感值）")
    return parser.parse_args()


def resolve_cookie_cache_path(cookie_cache_file: str | None = None) -> Path:
    """解析Cookie缓存路径；默认按OS用户隔离，避免多人共用目录时相互覆盖。"""
    if cookie_cache_file:
        return Path(cookie_cache_file).expanduser()

    env_value = os.getenv(DEFAULT_COOKIE_CACHE_ENV)
    if env_value:
        return Path(env_value).expanduser()

    return Path.home() / COOKIE_CACHE_FILE


def ensure_excel_writable(excel_path: str) -> None:
    """在执行创建/更新前检查Excel是否可写，避免处理完成后回写失败。"""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    try:
        with path.open("a+b") as handle:
            if os.name != "nt":
                return

            import msvcrt

            handle.seek(0)
            msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
            msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
    except PermissionError as e:
        raise PermissionError(f"Excel 文件被占用或无写权限: {excel_path}") from e
    except OSError as e:
        raise PermissionError(f"Excel 文件当前不可写（可能被占用）: {excel_path}") from e


def write_precheck_status_file(status_file: str, excel_path: str, ok: bool, message: str) -> None:
    status_payload = {
        "ok": ok,
        "excel": excel_path,
        "message": message,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    output_path = Path(status_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        json.dump(status_payload, handle, ensure_ascii=False, indent=2)


def load_json_file(filepath: str) -> dict[str, Any]:
    path = Path(filepath)
    if not path.exists():
        log.info(f"配置文件不存在，使用内置默认配置: {filepath}")
        return {}

    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def get_column_index(cell_reference: str) -> int:
    column_letters = ""
    for character in cell_reference:
        if character.isalpha():
            column_letters += character
        else:
            break

    index = 0
    for character in column_letters.upper():
        index = index * 26 + (ord(character) - ord("A") + 1)
    return index - 1


def column_index_to_letters(index: int) -> str:
    if index < 0:
        raise ValueError(f"列索引不能为负数: {index}")
    letters = ""
    current = index + 1
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def split_cell_reference(cell_reference: str) -> tuple[str, str]:
    letters = ""
    digits = ""
    for char in cell_reference:
        if char.isalpha():
            letters += char
        elif char.isdigit():
            digits += char
    return letters, digits


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    shared_strings_path = "xl/sharedStrings.xml"
    if shared_strings_path not in archive.namelist():
        return []

    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(archive.read(shared_strings_path))
    values: list[str] = []
    for item in root.findall("main:si", namespace):
        text_parts = [node.text or "" for node in item.findall(".//main:t", namespace)]
        values.append("".join(text_parts))
    return values


def load_workbook_sheet_map(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook_namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main", "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    rel_namespace = {"pkg": "http://schemas.openxmlformats.org/package/2006/relationships"}

    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    relationship_map: dict[str, str] = {}
    for relation in rel_root.findall("pkg:Relationship", rel_namespace):
        relation_id = relation.attrib.get("Id")
        target = relation.attrib.get("Target")
        if relation_id and target:
            normalized_target = target.lstrip("/")
            if not normalized_target.startswith("xl/"):
                normalized_target = f"xl/{normalized_target}"
            relationship_map[relation_id] = normalized_target

    sheet_map: dict[str, str] = {}
    for sheet in workbook_root.findall("main:sheets/main:sheet", workbook_namespace):
        name = sheet.attrib.get("name")
        relation_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if name and relation_id and relation_id in relationship_map:
            sheet_map[name] = relationship_map[relation_id]
    return sheet_map


def get_cell_text(cell: ET.Element, shared_strings: list[str], namespace: dict[str, str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        text_nodes = cell.findall(".//main:t", namespace)
        return "".join(node.text or "" for node in text_nodes)

    value_node = cell.find("main:v", namespace)
    raw_value = value_node.text if value_node is not None and value_node.text is not None else ""
    if cell_type == "s" and raw_value.isdigit():
        index = int(raw_value)
        return shared_strings[index] if index < len(shared_strings) else ""
    return raw_value


def find_cell_by_column(row_elem: ET.Element, column_letters: str, namespace: dict[str, str]) -> ET.Element | None:
    for cell in row_elem.findall("main:c", namespace):
        ref = cell.attrib.get("r", "")
        letters, _ = split_cell_reference(ref)
        if letters.upper() == column_letters.upper():
            return cell
    return None


def set_cell_inline_text(cell: ET.Element, row_number: str, column_letters: str, text: str, main_ns: str) -> None:
    cell.attrib["r"] = f"{column_letters}{row_number}"
    cell.attrib["t"] = "inlineStr"
    for child in list(cell):
        cell.remove(child)

    inline_elem = ET.SubElement(cell, f"{{{main_ns}}}is")
    text_elem = ET.SubElement(inline_elem, f"{{{main_ns}}}t")
    text_elem.text = text


def sort_row_cells(row_elem: ET.Element, namespace: dict[str, str]) -> None:
    cells = row_elem.findall("main:c", namespace)
    if len(cells) <= 1:
        return

    ordered_cells = sorted(cells, key=lambda c: get_column_index(c.attrib.get("r", "A1")))
    for cell in cells:
        row_elem.remove(cell)
    for cell in ordered_cells:
        row_elem.append(cell)


def shift_dimension_ref_for_inserted_first_column(dimension_ref: str) -> str:
    if not dimension_ref:
        return dimension_ref

    def _shift_ref(ref_text: str) -> str:
        letters, digits = split_cell_reference(ref_text)
        if not letters:
            return ref_text
        new_letters = column_index_to_letters(get_column_index(letters + (digits or "1")) + 1)
        return f"{new_letters}{digits}" if digits else new_letters

    if ":" in dimension_ref:
        start_ref, end_ref = dimension_ref.split(":", 1)
        return f"{_shift_ref(start_ref)}:{_shift_ref(end_ref)}"
    return _shift_ref(dimension_ref)


def write_ticket_ids_to_excel(
    excel_path: str,
    sheet_name: str | None,
    results: list[dict[str, Any]],
) -> None:
    successful_rows = {
        int(item["row"]): str(item["ticket_id"])
        for item in results
        if item.get("success") and item.get("ticket_id") not in (None, "", "DRY_RUN", "unknown")
    }
    if not successful_rows:
        return

    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError(f"缺少openpyxl依赖，无法回写Excel: {e}")

    workbook = load_workbook(excel_path)
    updated_count = 0
    try:
        target_sheet_name = sheet_name or workbook.sheetnames[0]
        if target_sheet_name not in workbook.sheetnames:
            log.warning(f"⚠️ 未能更新Excel中的ticket_id：未找到工作表 {target_sheet_name}")
            return

        worksheet = workbook[target_sheet_name]

        header_a_value = worksheet.cell(row=1, column=1).value
        has_ticket_id_column = str(header_a_value).strip().casefold() == "ticket_id" if header_a_value is not None else False
        if not has_ticket_id_column:
            worksheet.insert_cols(1)
            worksheet.cell(row=1, column=1).value = "ticket_id"

        data_row_numbers: list[int] = []
        for row_num in range(2, worksheet.max_row + 1):
            row_has_value = False
            for col_num in range(2, worksheet.max_column + 1):
                value = worksheet.cell(row=row_num, column=col_num).value
                if value not in (None, ""):
                    row_has_value = True
                    break
            if row_has_value:
                data_row_numbers.append(row_num)

        for result_row_index, ticket_id in successful_rows.items():
            if result_row_index < 0 or result_row_index >= len(data_row_numbers):
                continue
            row_num = data_row_numbers[result_row_index]
            worksheet.cell(row=row_num, column=1).value = ticket_id
            updated_count += 1

        if updated_count == 0:
            return

        try:
            workbook.save(excel_path)
            log.info(f"📝 已回写Excel ticket_id列: 更新 {updated_count} 条记录")
        except PermissionError:
            source_path = Path(excel_path)
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            fallback_path = source_path.with_name(f"{source_path.stem}.ticketid_updated_{timestamp}{source_path.suffix}")
            workbook.save(fallback_path)
            log.warning(
                f"⚠️ Excel文件被占用，已将ticket_id回写到新文件: {fallback_path}"
            )
    finally:
        workbook.close()


def parse_xlsx_rows(filepath: str, sheet_name: str | None = None) -> tuple[str, list[dict[str, str]]]:
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

    with zipfile.ZipFile(filepath) as archive:
        shared_strings = load_shared_strings(archive)
        sheet_map = load_workbook_sheet_map(archive)
        if not sheet_map:
            raise ValueError("Excel 中未找到工作表")

        target_sheet_name = sheet_name or next(iter(sheet_map.keys()))
        if target_sheet_name not in sheet_map:
            raise ValueError(f"未找到工作表: {target_sheet_name}")

        worksheet_path = sheet_map[target_sheet_name]
        worksheet_root = ET.fromstring(archive.read(worksheet_path))

        # 解析当前sheet的超链接关系，避免 ticket_id 为链接时被误判为空
        hyperlink_targets: dict[str, str] = {}
        sheet_rels_path = ""
        if worksheet_path.startswith("xl/"):
            rel_name = Path(worksheet_path).name + ".rels"
            sheet_rels_path = str(Path(worksheet_path).parent / "_rels" / rel_name).replace("\\", "/")

        rel_map: dict[str, str] = {}
        if sheet_rels_path and sheet_rels_path in archive.namelist():
            rel_root = ET.fromstring(archive.read(sheet_rels_path))
            for rel_elem in rel_root.findall("rel:Relationship", rel_ns):
                rel_id = rel_elem.attrib.get("Id", "")
                target = rel_elem.attrib.get("Target", "")
                if rel_id and target:
                    rel_map[rel_id] = target

        hyperlink_elems = worksheet_root.findall("main:hyperlinks/main:hyperlink", namespace)
        for h_elem in hyperlink_elems:
            ref_text = (h_elem.attrib.get("ref") or "").strip()
            rel_id = h_elem.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
            location = (h_elem.attrib.get("location") or "").strip()
            display = (h_elem.attrib.get("display") or "").strip()

            if not ref_text:
                continue
            # 处理类似 A2:B2 的范围，按左上角单元格定位
            cell_ref = ref_text.split(":", 1)[0].strip()
            target_text = display or location
            if not target_text and rel_id:
                target_text = rel_map.get(rel_id, "")
            if target_text:
                hyperlink_targets[cell_ref] = target_text

        sheet_rows = worksheet_root.findall("main:sheetData/main:row", namespace)
        parsed_rows: list[list[str]] = []
        max_columns = 0

        for row in sheet_rows:
            values_by_index: dict[int, str] = {}
            for cell in row.findall("main:c", namespace):
                cell_reference = cell.attrib.get("r", "")
                column_index = get_column_index(cell_reference) if cell_reference else len(values_by_index)
                cell_type = cell.attrib.get("t")
                raw_value = ""

                if cell_type == "inlineStr":
                    text_nodes = cell.findall(".//main:t", namespace)
                    raw_value = "".join(node.text or "" for node in text_nodes)
                else:
                    value_node = cell.find("main:v", namespace)
                    raw_value = value_node.text if value_node is not None and value_node.text is not None else ""
                    if not raw_value:
                        formula_node = cell.find("main:f", namespace)
                        if formula_node is not None and formula_node.text:
                            raw_value = f"={formula_node.text.strip()}"
                    if cell_type == "s" and raw_value.isdigit():
                        string_index = int(raw_value)
                        raw_value = shared_strings[string_index] if string_index < len(shared_strings) else ""

                if not raw_value and cell_reference in hyperlink_targets:
                    raw_value = hyperlink_targets[cell_reference]

                values_by_index[column_index] = raw_value.strip()

            if values_by_index:
                row_values = [""] * (max(values_by_index.keys()) + 1)
                for column_index, value in values_by_index.items():
                    row_values[column_index] = value
                max_columns = max(max_columns, len(row_values))
                parsed_rows.append(row_values)

        if not parsed_rows:
            raise ValueError("Excel 为空，未读取到任何数据")

        normalized_rows = [row + ([""] * (max_columns - len(row))) for row in parsed_rows]
        headers = [value.strip() for value in normalized_rows[0]]
        if not any(headers):
            raise ValueError("Excel 表头为空")

        records: list[dict[str, str]] = []
        for row_values in normalized_rows[1:]:
            row_dict: dict[str, str] = {}
            has_value = False
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row_values[index].strip() if index < len(row_values) else ""
                if value:
                    has_value = True
                row_dict[header] = value
            if has_value:
                records.append(row_dict)

        return target_sheet_name, records


def load_excel_rows(filepath: str, sheet_name: str | None = None) -> list[dict[str, str]]:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {filepath}")

    file_extension = path.suffix.lower()
    if file_extension != ".xlsx":
        raise ValueError(f"当前仅支持 .xlsx 文件，实际文件类型为: {file_extension}")

    loaded_sheet_name, records = parse_xlsx_rows(filepath, sheet_name)
    log.info(f"读取 Excel: {filepath}  工作表: {loaded_sheet_name}  共 {len(records)} 条记录")
    return records


def get_row_value(row: dict[str, str], column_name: str) -> str:
    target_name = column_name.casefold()
    for key, value in row.items():
        if key.casefold() == target_name:
            return value
    return ""


def validate_columns(rows: list[dict[str, str]]) -> None:
    if not rows:
        raise ValueError("Excel 中没有可用数据行")

    missing = [column for column in REQUIRED_EXCEL_COLUMNS if get_row_value(rows[0], column) == ""]
    if missing:
        raise ValueError(f"Excel 缺少必填列: {missing}")


def validate_ticket_id_values(rows: list[dict[str, str]], start_row_offset: int = 0) -> None:
    """校验 ticket_id 列：有内容时必须为纯数字；超链接/公式/其他文本均视为不合规。"""
    issues: list[str] = []

    for index, row in enumerate(rows):
        ticket_id_raw = get_row_value(row, "ticket_id").strip()
        if ticket_id_raw == "":
            continue

        excel_row_number = start_row_offset + index + 2

        # 先判有内容，再判合规（仅允许纯数字ID）
        if ticket_id_raw.startswith("="):
            issues.append(f"Excel第{excel_row_number}行: ticket_id 为公式，需改为纯数字ID或留空")
        elif ticket_id_raw.lower().startswith("http://") or ticket_id_raw.lower().startswith("https://"):
            issues.append(f"Excel第{excel_row_number}行: ticket_id 为链接，需改为纯数字ID或留空")
        elif not ticket_id_raw.isdigit():
            issues.append(f"Excel第{excel_row_number}行: ticket_id='{ticket_id_raw}' 非纯数字，需修正")

    if issues:
        preview = "\n".join(issues[:10])
        suffix = "\n..." if len(issues) > 10 else ""
        raise ValueError(
            "检测到 ticket_id 列存在不合规内容，请先修正后再执行。\n"
            "规则：ticket_id 有内容时必须是纯数字；如果是超链接或公式，请改成纯数字ID文本。\n"
            f"明细:\n{preview}{suffix}"
        )


def get_cached_cookie(cache_path: Path) -> str | None:
    """读取本地缓存的Cookie"""
    if cache_path.exists():
        try:
            with cache_path.open("r", encoding="utf-8") as f:
                cookie = f.read().strip()
                if cookie:
                    log.info(f"📂 使用本地缓存的Cookie: {cache_path}")
                    return cookie
        except Exception as e:
            log.debug(f"读取Cookie缓存失败: {e}")
    return None


def save_cookie_to_cache(cookie: str, cache_path: Path) -> None:
    """保存有效的Cookie到本地缓存"""
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(normalize_cookie_string(cookie), encoding="utf-8")
        cache_path.chmod(0o600)  # 仅限所有者读写
        log.info(f"💾 Cookie已保存到本地缓存: {cache_path}")
    except Exception as e:
        log.warning(f"保存Cookie缓存失败: {e}")


def normalize_cookie_string(cookie: str) -> str:
    """标准化Cookie字符串，兼容从DevTools复制的不同格式。"""
    normalized = (cookie or "").strip()
    if not normalized:
        return ""

    # 兼容 "Cookie: a=b; c=d" 形式
    if normalized.lower().startswith("cookie:"):
        normalized = normalized.split(":", 1)[1].strip()

    # 去掉包裹引号
    if (normalized.startswith('"') and normalized.endswith('"')) or (
        normalized.startswith("'") and normalized.endswith("'")
    ):
        normalized = normalized[1:-1].strip()

    # 去掉Set-Cookie属性，避免把Path/Domain/HttpOnly等带入请求头
    cookie_parts = []
    attr_keys = {
        "path",
        "domain",
        "expires",
        "max-age",
        "secure",
        "httponly",
        "samesite",
        "priority",
        "partitioned",
    }
    for part in normalized.split(";"):
        piece = part.strip()
        if not piece:
            continue
        key = piece.split("=", 1)[0].strip().lower()
        if key in attr_keys:
            continue
        if "=" in piece:
            cookie_parts.append(piece)

    normalized_cookie = "; ".join(cookie_parts) if cookie_parts else normalized

    # 粘贴整段浏览器Cookie时，优先保留认证关键Cookie，避免头部过大被服务端拒绝。
    auth_cookie_keys = {
        "mysapsso2",
        "mod_auth_openidc_session",
        "idsid",
        "es_mod_auth_flag",
        "ighfloggedin",
    }
    parsed_pairs = []
    for piece in normalized_cookie.split(";"):
        item = piece.strip()
        if not item or "=" not in item:
            continue
        key, value = item.split("=", 1)
        parsed_pairs.append((key.strip(), value.strip()))

    selected = [f"{k}={v}" for k, v in parsed_pairs if k.lower() in auth_cookie_keys]
    if selected:
        return "; ".join(selected)

    return normalized_cookie


def verify_cookie_valid(
    cookie: str,
    api_url: str,
    timeout: int,
    insecure: bool,
    extra_headers: dict[str, str] | None = None,
) -> bool:
    """验证Cookie是否仍然有效"""
    try:
        normalized_cookie = normalize_cookie_string(cookie)
        if not normalized_cookie:
            return False

        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "Cookie": normalized_cookie,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        }
        if extra_headers:
            headers.update(extra_headers)
        
        test_payload = {
            "requests": [
                {
                    "api_client": DEFAULT_API_CLIENT,
                    "tran_id": str(uuid.uuid4()).upper(),
                    "command": "lookup_options",
                    "command_args": {
                        "tenant": "server_platf",
                        "subject": "test_case",
                    },
                    "var_args": [],
                    "copy_args": [],
                }
            ]
        }
        req = request.Request(
            api_url,
            data=json.dumps(test_payload).encode("utf-8"),
            headers=headers,
            method="POST",
        )
        
        ctx = create_ssl_context(insecure)
        with request.urlopen(req, timeout=timeout, context=ctx) as resp:
            response_text = resp.read().decode("utf-8")
            response_data = json.loads(response_text)
            
            # 检查是否有错误响应（如认证失败）
            if isinstance(response_data, dict) and response_data.get("responses"):
                first_resp = response_data["responses"][0] if response_data["responses"] else {}
                if first_resp.get("status") == "error":
                    error_msg = str(first_resp.get("errordetail", ""))
                    err_lower = error_msg.lower()
                    if (
                        "authent" in err_lower
                        or "invalid" in err_lower
                        or "login" in err_lower
                        or "forbidden" in err_lower
                        or "permission" in err_lower
                        or "session" in err_lower
                    ):
                        return False
            return True
    except error.HTTPError as e:
        try:
            response_text = e.read().decode("utf-8", errors="replace")
        except Exception:
            response_text = ""
        lower_text = response_text.lower()
        if e.code in (401, 403) or any(
            keyword in lower_text
            for keyword in ("auth", "login", "forbidden", "invalid", "session")
        ):
            return False
        log.warning(f"⚠️ Cookie校验遇到非认证HTTP错误(HTTP {e.code})，将继续尝试使用该Cookie")
        return True
    except Exception as e:
        log.warning(f"⚠️ Cookie校验异常({e})，将继续尝试使用该Cookie")
        return True


def login_with_credentials(
    username: str,
    password: str,
    timeout: int,
    insecure: bool,
    auth_debug: bool = False,
) -> str | None:
    """使用账户密码登录HSDES并获取新Cookie"""
    try:
        log.info(f"🔐 尝试使用账户 {username} 登录 HSDES...")

        credentials = base64.b64encode(f"{username}:{password}".encode()).decode("ascii")
        ctx = create_ssl_context(insecure)

        cookie_jar = http.cookiejar.CookieJar()
        handlers: list[Any] = [request.HTTPCookieProcessor(cookie_jar)]
        if ctx:
            handlers.append(request.HTTPSHandler(context=ctx))
        opener = request.build_opener(*handlers)

        base_headers = {
            "Authorization": f"Basic {credentials}",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }

        # 依次访问登录页与主页，尽量覆盖SSO跳转链并收集Cookie。
        try:
            for idx, url in enumerate((HSDES_LOGIN_URL, HSDES_HOME_URL), start=1):
                req = request.Request(url, headers=base_headers)
                with opener.open(req, timeout=timeout) as resp:
                    log.info(f"✅ 登录步骤{idx}成功，状态码: {resp.status}")
                    if auth_debug:
                        log.info(f"[AUTH-DEBUG] step={idx} final_url={resp.geturl()}")

            cookies = [
                c for c in cookie_jar
                if (c.domain and "hsdes.intel.com" in c.domain.lower()) or c.domain == ""
            ]
            if auth_debug:
                cookie_names = ", ".join(sorted({c.name for c in cookies})) if cookies else "<none>"
                log.info(f"[AUTH-DEBUG] captured_cookie_count={len(cookies)} names={cookie_names}")

            if cookies:
                cookie_str = "; ".join(f"{c.name}={c.value}" for c in cookies if c.name and c.value)
                cookie_str = normalize_cookie_string(cookie_str)
                if cookie_str:
                    log.info("🍪 已获取新的Session Cookie")
                    return cookie_str

            log.warning("⚠️  登录成功但未获得Cookie")
            return None

        except error.HTTPError as e:
            if e.code == 401 or e.code == 403:
                log.error(f"❌ 登录失败: 账户或密码错误 (HTTP {e.code})")
            else:
                log.error(f"❌ 登录失败: HTTP {e.code}")
            if auth_debug:
                try:
                    body = e.read().decode("utf-8", errors="replace")
                    log.info(f"[AUTH-DEBUG] http_error_body_snippet={body[:240].replace(chr(10), ' ')}")
                except Exception:
                    pass
            return None

    except Exception as e:
        log.error(f"❌ 登录异常: {e}")
        return None


def auto_login_hsdes(timeout: int, insecure: bool) -> str | None:
    """尝试使用Windows集成认证自动登录HSDES"""
    try:
        log.info("🔑 尝试Windows集成认证自动登录...")
        
        # 尝试使用系统认证（适用于域环境）
        ctx = create_ssl_context(insecure)
        req = request.Request(HSDES_HOME_URL)
        
        with request.urlopen(req, timeout=timeout, context=ctx) as resp:
            # 获取Set-Cookie header中的cookie
            for header, value in resp.headers.items():
                if header.lower() == "set-cookie":
                    log.info("✅ 自动登录成功")
                    return value.split(";")[0] if value else None
        
        log.warning("⚠️  自动登录未获得Cookie")
        return None
        
    except Exception as e:
        log.debug(f"自动登录失败: {e}")
        return None


def get_or_login_cookie(
    provided_cookie: str | None,
    api_url: str,
    timeout: int,
    insecure: bool,
    cookie_cache_path: Path,
    use_cookie_cache: bool = True,
    extra_headers: dict[str, str] | None = None,
    allow_interactive: bool = True,
    username: str | None = None,
    password: str | None = None,
    auth_debug: bool = False,
) -> str | None:
    """
    获取有效的Cookie，尝试多种方式：
    1. 使用提供的Cookie（如果有效）
    2. 使用本地缓存的Cookie（如果有效）
    3. 提示用户输入账户密码进行登录（实验性）
    4. 提示用户手动提供Cookie
    """
    def prompt_manual_cookie() -> str | None:
        log.info("\n" + "="*60)
        log.info("如何获取Cookie:")
        log.info("1. 打开浏览器访问 https://hsdes.intel.com/appstore/article-one/")
        log.info("2. 按 F12 打开DevTools")
        log.info("3. 选择 Application → Cookies")
        log.info("4. 复制Cookie字符串")
        log.info("5. 粘贴到下面的输入框中")
        log.info("="*60 + "\n")

        raw_cookie = input("请粘贴Cookie字符串（可用Ctrl+V）: ").strip()
        cookie = normalize_cookie_string(raw_cookie)
        if not cookie:
            log.error("❌ Cookie不能为空")
            return None

        if verify_cookie_valid(cookie, api_url, timeout, insecure, extra_headers=extra_headers):
            if use_cookie_cache:
                save_cookie_to_cache(cookie, cookie_cache_path)
                log.info("✅ Cookie已保存，下次执行时会自动使用")
            else:
                log.info("✅ Cookie验证通过（当前运行未启用缓存）")
            return cookie

        log.error("❌ Cookie无效或已过期，请确认已完整复制并重新获取")
        return None

    # 方案1: 检查提供的Cookie
    provided_cookie = normalize_cookie_string(provided_cookie or "")
    if provided_cookie:
        if verify_cookie_valid(provided_cookie, api_url, timeout, insecure, extra_headers=extra_headers):
            log.info("✓ 提供的Cookie有效")
            if use_cookie_cache:
                save_cookie_to_cache(provided_cookie, cookie_cache_path)
            return provided_cookie
        else:
            log.warning("✗ 提供的Cookie已过期或无效")
    
    # 方案2: 检查缓存的Cookie
    cached_cookie = get_cached_cookie(cookie_cache_path) if use_cookie_cache else None
    if cached_cookie:
        if verify_cookie_valid(cached_cookie, api_url, timeout, insecure, extra_headers=extra_headers):
            log.info("✓ 缓存的Cookie有效")
            return cached_cookie
        else:
            log.warning("✗ 缓存的Cookie已过期，尝试重新登录...")
            try:
                cookie_cache_path.unlink(missing_ok=True)
            except Exception:
                pass
    
    # 方案3: 尝试自动登录（实验性，仅当提供了凭证）
    if username and password:
        log.info(f"🔐 尝试使用账户 {username} 登录（实验功能）")
        new_cookie = login_with_credentials(username, password, timeout, insecure, auth_debug=auth_debug)
        if new_cookie and verify_cookie_valid(new_cookie, api_url, timeout, insecure, extra_headers=extra_headers):
            if use_cookie_cache:
                save_cookie_to_cache(new_cookie, cookie_cache_path)
            return new_cookie
        else:
            log.warning("⚠️  自动登录未能获取有效Cookie，请手动提供")
    
    # 方案4: 交互式提示
    if allow_interactive:
        log.info("\n" + "="*60)
        log.info("需要登录HSDES - 请选择认证方式")
        log.info("="*60)
        log.info("方式1: 使用账户密码登录（实验功能，可能不可用）")
        log.info("方式2: 手动提供Cookie（推荐）")
        log.info("="*60)
        
        choice = input("请选择 (1/2，默认2): ").strip() or "2"
        
        if choice == "1":
            username = input("请输入Intel账户（用户名）: ").strip()
            if not username:
                log.error("❌ 用户名不能为空")
                return None
            
            password = getpass.getpass("请输入密码（不会显示）: ")
            if not password:
                log.error("❌ 密码不能为空")
                return None
            
            new_cookie = login_with_credentials(username, password, timeout, insecure, auth_debug=auth_debug)
            if new_cookie and verify_cookie_valid(new_cookie, api_url, timeout, insecure, extra_headers=extra_headers):
                if use_cookie_cache:
                    save_cookie_to_cache(new_cookie, cookie_cache_path)
                return new_cookie
            else:
                log.warning("⚠️  登录未成功，自动切换到方式2（手动提供Cookie）")
                return prompt_manual_cookie()
        
        elif choice == "2":
            return prompt_manual_cookie()

        else:
            log.warning("⚠️  无效选择，自动切换到方式2（手动提供Cookie）")
            return prompt_manual_cookie()
    
    log.error("❌ 无法获取有效的Cookie")
    return None


def build_runtime_config(args: argparse.Namespace, file_config: dict[str, Any]) -> dict[str, Any]:
    field_mapping = dict(DEFAULT_FIELD_MAPPING)
    field_mapping.update(file_config.get("field_mapping", {}))

    static_fields = dict(DEFAULT_STATIC_FIELDS)
    static_fields.update(file_config.get("static_fields", {}))
    if args.subject:
        static_fields["subject"] = args.subject

    api_url = args.api_url or file_config.get("api_url")
    auth_header = args.auth_header or file_config.get("auth_header", "Authorization")
    auth_scheme = args.auth_scheme if args.auth_scheme is not None else file_config.get("auth_scheme", "Bearer")
    token_env = args.token_env or file_config.get("token_env", DEFAULT_TOKEN_ENV)
    token = args.token or os.getenv(token_env) or file_config.get("token")
    cookie_env = args.cookie_env or file_config.get("cookie_env", DEFAULT_COOKIE_ENV)
    cookie = args.cookie or os.getenv(cookie_env) or file_config.get("cookie")
    cookie_cache_file = args.cookie_cache_file or file_config.get("cookie_cache_file")
    use_cookie_cache = not args.no_cookie_cache
    extra_headers = dict(file_config.get("extra_headers", {}))
    envelope = dict(file_config.get("envelope", {}))
    response_path = list(file_config.get("response_path", []))

    return {
        "api_url": api_url,
        "auth_header": auth_header,
        "auth_scheme": auth_scheme,
        "token_env": token_env,
        "token": token,
        "cookie_env": cookie_env,
        "cookie": cookie,
        "cookie_cache_path": resolve_cookie_cache_path(cookie_cache_file),
        "use_cookie_cache": use_cookie_cache,
        "field_mapping": field_mapping,
        "static_fields": static_fields,
        "extra_headers": extra_headers,
        "envelope": envelope,
        "response_path": response_path,
    }


def format_field_value(api_field: str, value: str) -> str:
    normalized_value = value.replace("\r\n", "\n").replace("\r", "\n")
    if api_field in MULTILINE_DISPLAY_FIELDS:
        lines = normalized_value.split("\n")
        return "<br/>".join(lines)
    return normalized_value


def build_payload(row: dict[str, str], field_mapping: dict[str, str], static_fields: dict[str, Any]) -> dict[str, Any]:
    payload = dict(static_fields)

    tenant_subject = payload.get("tenant_subject", "")
    if isinstance(tenant_subject, str) and "." in tenant_subject:
        payload["subject"] = tenant_subject.split(".", 1)[1]

    for excel_column, api_field in field_mapping.items():
        value = get_row_value(row, excel_column)
        if value != "":
            payload[api_field] = format_field_value(api_field, value)
    return payload


def build_headers(
    auth_header: str,
    auth_scheme: str | None,
    token: str | None,
    cookie: str | None,
    extra_headers: dict[str, str],
) -> dict[str, str]:
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    headers.update(extra_headers)

    if token:
        auth_value = token if not auth_scheme else f"{auth_scheme} {token}"
        headers[auth_header] = auth_value
    if cookie:
        # Accept copied values like "cookie\nkey1=...; key2=..." from devtools.
        normalized_cookie = " ".join(cookie.splitlines()).strip()
        lowered = normalized_cookie.lower()
        if lowered.startswith("cookie"):
            if ":" in normalized_cookie[:16]:
                normalized_cookie = normalized_cookie.split(":", 1)[1].strip()
            else:
                normalized_cookie = normalized_cookie[len("cookie"):].strip()
        headers["Cookie"] = normalized_cookie

    return headers


def build_enveloped_payload(payload: dict[str, Any], envelope: dict[str, Any]) -> dict[str, Any]:
    if not envelope:
        return payload

    command_args = dict(envelope.get("command_args", {}))
    payload_target = str(envelope.get("payload_target", "command_args")).strip().lower()

    base_var_args = list(envelope.get("var_args", []))
    base_copy_args = list(envelope.get("copy_args", []))

    dynamic_var_args = [{key: value} for key, value in payload.items()]

    if payload_target == "command_args":
        command_args.update(payload)
        var_args = base_var_args
    elif payload_target == "var_args":
        var_args = base_var_args + dynamic_var_args
    else:
        raise ValueError(f"不支持的 payload_target: {payload_target}，仅支持 command_args 或 var_args")

    request_payload = {
        "api_client": envelope.get("api_client", DEFAULT_API_CLIENT),
        "tran_id": str(uuid.uuid4()).upper(),
        "command": envelope.get("command", "create_record"),
        "command_args": command_args,
        "var_args": var_args,
        "copy_args": base_copy_args,
    }
    return {"requests": [request_payload]}


def create_ssl_context(insecure: bool) -> ssl.SSLContext | None:
    if not insecure:
        return None
    return ssl._create_unverified_context()


def post_ticket(
    api_url: str,
    payload: dict[str, Any],
    headers: dict[str, str],
    timeout: int,
    insecure: bool,
) -> tuple[bool, dict[str, Any], str]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request_obj = request.Request(api_url, data=body, headers=headers, method="POST")
    context = create_ssl_context(insecure)

    try:
        with request.urlopen(request_obj, timeout=timeout, context=context) as response:
            response_text = response.read().decode("utf-8", errors="replace")
            response_json = json.loads(response_text) if response_text else {}
            return True, response_json, ""
    except error.HTTPError as exc:
        response_text = exc.read().decode("utf-8", errors="replace")
        return False, {}, f"HTTP {exc.code}: {response_text}"
    except error.URLError as exc:
        return False, {}, f"网络错误: {exc.reason}"
    except json.JSONDecodeError as exc:
        return False, {}, f"响应不是合法 JSON: {exc}"


def update_ticket_fields(
    api_url: str,
    headers: dict[str, str],
    timeout: int,
    insecure: bool,
    tenant: str,
    subject: str,
    ticket_id: str,
    update_fields: dict[str, str],
) -> tuple[bool, dict[str, Any], str]:
    """在创建后更新特定字段（例如create阶段不会落库的free_tag字段）。"""
    if not update_fields:
        return True, {}, ""

    payload = {
        "requests": [
            {
                "api_client": DEFAULT_API_CLIENT,
                "tran_id": str(uuid.uuid4()).upper(),
                "command": "update_record_with_fetch",
                "command_args": {
                    "tenant": tenant,
                    "subject": subject,
                    "id": str(ticket_id),
                },
                "var_args": [{k: v} for k, v in update_fields.items()],
                "copy_args": [],
            }
        ]
    }
    return post_ticket(api_url, payload, headers, timeout, insecure)


def fetch_ticket_fields_by_id(
    api_url: str,
    headers: dict[str, str],
    timeout: int,
    insecure: bool,
    tenant: str,
    subject: str,
    ticket_id: str,
) -> tuple[bool, dict[str, Any], str]:
    """读取指定ticket当前字段。当前ESService无独立读命令，使用空更新请求进行只读式获取。"""
    payload = {
        "requests": [
            {
                "api_client": DEFAULT_API_CLIENT,
                "tran_id": str(uuid.uuid4()).upper(),
                "command": "update_record_with_fetch",
                "command_args": {
                    "tenant": tenant,
                    "subject": subject,
                    "id": str(ticket_id),
                },
                "var_args": [],
                "copy_args": [],
            }
        ]
    }
    return post_ticket(api_url, payload, headers, timeout, insecure)


def normalize_value_for_compare(field: str, value: Any) -> str:
    text = "" if value is None else str(value)
    normalized = text.strip()
    if field == "notify":
        normalized = ",".join(part.strip() for part in normalized.split(",") if part.strip())
    if "description" in field:
        normalized = normalized.replace("<br />", "<br/>").replace("<br>", "<br/>")
    return normalized


def diff_ticket_fields(desired_fields: dict[str, Any], current_fields: dict[str, Any]) -> dict[str, str]:
    updates: dict[str, str] = {}
    for key, desired_value in desired_fields.items():
        desired_norm = normalize_value_for_compare(key, desired_value)
        current_norm = normalize_value_for_compare(key, current_fields.get(key, ""))
        if desired_norm != current_norm:
            updates[key] = str(desired_value)
    return updates


def extract_ticket_id(response_json: dict[str, Any]) -> str:
    candidate_paths = [
        ("responses", "0", "result_table", "0", "id"),
        ("responses", "0", "result_table", "0", "from_id"),
        ("responses", "0", "result_table", "0", "article_id"),
        ("responses", "0", "id"),
        ("id",),
        ("ticket_id",),
        ("data", "id"),
        ("data", "ticket_id"),
        ("result", "id"),
        ("result", "ticket_id"),
    ]

    for path in candidate_paths:
        current: Any = response_json
        found = True
        for key in path:
            if isinstance(current, list):
                if not key.isdigit():
                    found = False
                    break
                index = int(key)
                if index >= len(current):
                    found = False
                    break
                current = current[index]
                continue
            if not isinstance(current, dict) or key not in current:
                found = False
                break
            current = current[key]
        if found and current not in (None, ""):
            return str(current)
    return "unknown"


def extract_by_path(response_json: dict[str, Any], response_path: list[str]) -> Any:
    current: Any = response_json
    for key in response_path:
        if isinstance(current, list):
            if not key.isdigit():
                return None
            index = int(key)
            if index >= len(current):
                return None
            current = current[index]
            continue
        if not isinstance(current, dict) or key not in current:
            return None
        current = current[key]
    return current


def extract_esservice_error(response_json: dict[str, Any]) -> str:
    responses = response_json.get("responses") if isinstance(response_json, dict) else None
    if not isinstance(responses, list) or not responses:
        return ""

    first = responses[0]
    if not isinstance(first, dict):
        return ""

    status = str(first.get("status", "")).lower()
    if status in ("", "success"):
        return ""

    messages = first.get("messages", [])
    if isinstance(messages, list) and messages:
        details: list[str] = []
        for item in messages:
            if isinstance(item, dict):
                summary = str(item.get("errorsummary", "")).strip()
                detail = str(item.get("errordetail", "")).strip()
                if summary and detail:
                    details.append(f"{summary}: {detail}")
                elif summary:
                    details.append(summary)
                elif detail:
                    details.append(detail)
            else:
                details.append(str(item))
        if details:
            return " | ".join(details)

    return f"ESService 返回状态: {status}"


def create_ticket(
    row: dict[str, str],
    api_url: str,
    headers: dict[str, str],
    field_mapping: dict[str, str],
    static_fields: dict[str, Any],
    envelope: dict[str, Any],
    response_path: list[str],
    timeout: int,
    insecure: bool,
    dry_run: bool,
) -> dict[str, Any]:
    title = get_row_value(row, "Title") or "(无标题)"
    raw_payload = build_payload(row, field_mapping, static_fields)
    payload = build_enveloped_payload(raw_payload, envelope)
    log.info(f"→ 创建 ticket: {title}")

    if dry_run:
        log.info(f"  [DRY RUN] payload = {json.dumps(payload, ensure_ascii=False)}")
        return {
            "success": True,
            "ticket_id": "DRY_RUN",
            "error": "",
            "payload": payload,
            "action": "created",
        }

    success, response_json, error_message = post_ticket(api_url, payload, headers, timeout, insecure)
    if not success:
        log.error(f"  ✗ 创建失败: {error_message}")
        return {
            "success": False,
            "ticket_id": "",
            "error": error_message,
            "payload": payload,
            "action": "create_failed",
        }

    esservice_error = extract_esservice_error(response_json)
    if esservice_error:
        log.error(f"  ✗ 创建失败: {esservice_error}")
        return {
            "success": False,
            "ticket_id": "",
            "error": esservice_error,
            "payload": payload,
            "response": response_json,
            "action": "create_failed",
        }

    extracted_value = extract_by_path(response_json, response_path) if response_path else None
    ticket_id = str(extracted_value) if extracted_value not in (None, "") else extract_ticket_id(response_json)

    # 部分字段（例如 free_tag_1/2/3）在 create 阶段被服务端忽略，需要创建后补一次 update。
    deferred_updates = {
        key: str(value)
        for key, value in raw_payload.items()
        if key in DEFERRED_UPDATE_FIELDS and value not in (None, "")
    }
    if deferred_updates:
        tenant = str(envelope.get("command_args", {}).get("tenant", "server_platf"))
        subject = str(envelope.get("command_args", {}).get("subject", "test_case"))
        updated, update_response_json, update_error = update_ticket_fields(
            api_url=api_url,
            headers=headers,
            timeout=timeout,
            insecure=insecure,
            tenant=tenant,
            subject=subject,
            ticket_id=ticket_id,
            update_fields=deferred_updates,
        )
        if not updated:
            log.warning(f"  ⚠️ 创建成功但补充更新字段失败: {update_error}")
        else:
            update_esservice_error = extract_esservice_error(update_response_json)
            if update_esservice_error:
                log.warning(f"  ⚠️ 创建成功但补充更新字段失败: {update_esservice_error}")
            else:
                log.info("  ✓ 已补充更新 free_tag 字段")

    log.info(f"  ✅ 创建成功  ticket_id={ticket_id}")
    return {
        "success": True,
        "ticket_id": ticket_id,
        "error": "",
        "payload": payload,
        "response": response_json,
        "action": "created",
    }


def save_results(excel_path: str, results: list[dict[str, Any]]) -> None:
    result_file = Path(excel_path).parent / "hsdes_create_results.json"
    with result_file.open("w", encoding="utf-8") as handle:
        json.dump(results, handle, ensure_ascii=False, indent=2)
    log.info(f"结果已保存: {result_file}")


def run_first_run_wizard(args: argparse.Namespace, runtime_config: dict[str, Any], row_count: int) -> None:
    """首次运行向导：展示关键配置并让用户选择执行策略。"""
    log.info("\n" + "=" * 60)
    log.info("🧭 首次运行向导")
    log.info("=" * 60)
    log.info(f"Excel: {args.excel}")
    log.info(f"记录数: {row_count}")
    log.info(f"API URL: {runtime_config.get('api_url') or '(未配置)'}")
    if runtime_config.get("use_cookie_cache"):
        log.info(f"Cookie缓存: 启用 ({runtime_config.get('cookie_cache_path')})")
    else:
        log.info("Cookie缓存: 禁用（仅本次运行）")

    if args.dry_run:
        log.info("当前已是 dry-run 模式，将继续执行预览。")
        return

    log.info("\n请选择本次执行方式:")
    log.info("1. 先 dry-run 预览 payload 后退出")
    log.info("2. 直接正式创建 ticket")
    log.info("3. 退出")
    choice = input("请输入选项 (1/2/3，默认2): ").strip() or "2"

    if choice == "1":
        args.dry_run = True
        log.info("已切换为 dry-run 模式。")
        return
    if choice == "2":
        log.info("继续正式创建流程。")
        return

    log.info("已取消本次执行。")
    sys.exit(0)


def main() -> None:
    args = parse_args()
    precheck_status_file = args.precheck_status_file or DEFAULT_PRECHECK_STATUS_FILE
    file_config = load_json_file(args.config)
    runtime_config = build_runtime_config(args, file_config)

    rows = load_excel_rows(args.excel, args.sheet)
    validate_columns(rows)
    rows = rows[args.start_row:]
    validate_ticket_id_values(rows, start_row_offset=args.start_row)

    if args.wizard:
        run_first_run_wizard(args, runtime_config, len(rows))

    if not args.dry_run:
        try:
            ensure_excel_writable(args.excel)
            log.info("✓ Excel写回预检查通过")
            write_precheck_status_file(
                status_file=precheck_status_file,
                excel_path=args.excel,
                ok=True,
                message="excel_writable",
            )
        except PermissionError:
            error_message = "excel_locked_or_not_writable"
            write_precheck_status_file(
                status_file=precheck_status_file,
                excel_path=args.excel,
                ok=False,
                message=error_message,
            )
            log.error(
                f"❌ Excel文件当前被占用或不可写: {args.excel}\n"
                "请先关闭该文件（包括Excel预览窗口/同步工具占用）后重试。"
            )
            log.error(f"预检查状态文件: {precheck_status_file}")
            sys.exit(2)

    if not args.dry_run and not runtime_config["api_url"]:
        raise ValueError("未提供 API URL。请通过 --api-url 或配置文件中的 api_url 指定。")
    
    # 方案B：自动登录或获取有效的Cookie
    if not args.dry_run:
        if runtime_config["token"]:
            log.info("✓ 使用提供的 API token 进行认证")
        else:
            if runtime_config["use_cookie_cache"]:
                log.info(f"ℹ️ Cookie缓存路径: {runtime_config['cookie_cache_path']}")
            else:
                log.info("ℹ️ 已禁用Cookie缓存（仅本次运行有效）")

            # 尝试获取或登录新的Cookie
            cookie = get_or_login_cookie(
                provided_cookie=runtime_config["cookie"],
                api_url=runtime_config["api_url"],
                timeout=args.timeout,
                insecure=args.insecure,
                cookie_cache_path=runtime_config["cookie_cache_path"],
                use_cookie_cache=runtime_config["use_cookie_cache"],
                extra_headers=runtime_config["extra_headers"],
                allow_interactive=True,
                username=args.username,
                password=args.password,
                auth_debug=args.auth_debug,
            )
            
            if not cookie:
                raise ValueError(
                    "❌ 无法获取有效的认证信息。请通过以下方式之一提供：\n"
                    "  1. --cookie 参数直接提供有效的Cookie\n"
                    "  2. 设置环境变量 HSDES_COOKIE\n"
                    "  3. 提供 --username 和 --password 进行自动登录\n"
                    "  4. 允许脚本交互式登录时输入账户密码"
                )
            
            # 更新runtime_config中的cookie
            runtime_config["cookie"] = cookie

    headers = build_headers(
        runtime_config["auth_header"],
        runtime_config["auth_scheme"],
        runtime_config["token"],
        runtime_config["cookie"],
        runtime_config["extra_headers"],
    )

    tenant = str(runtime_config["envelope"].get("command_args", {}).get("tenant", "server_platf"))
    subject = str(runtime_config["envelope"].get("command_args", {}).get("subject", "test_case"))

    results: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=args.start_row):
        row_ticket_id = get_row_value(row, "ticket_id")
        if not args.dry_run and row_ticket_id:
            log.info(f"→ 检查并按需更新 ticket: {row_ticket_id}")

            desired_fields = build_payload(
                row=row,
                field_mapping=runtime_config["field_mapping"],
                static_fields=runtime_config["static_fields"],
            )

            fetched, current_response, fetch_error = fetch_ticket_fields_by_id(
                api_url=runtime_config["api_url"],
                headers=headers,
                timeout=args.timeout,
                insecure=args.insecure,
                tenant=tenant,
                subject=subject,
                ticket_id=row_ticket_id,
            )

            if not fetched:
                result = {
                    "success": False,
                    "ticket_id": row_ticket_id,
                    "error": f"读取原ticket失败: {fetch_error}",
                    "payload": desired_fields,
                    "action": "update_failed",
                }
            else:
                fetch_esservice_error = extract_esservice_error(current_response)
                if fetch_esservice_error:
                    result = {
                        "success": False,
                        "ticket_id": row_ticket_id,
                        "error": f"读取原ticket失败: {fetch_esservice_error}",
                        "payload": desired_fields,
                        "response": current_response,
                        "action": "update_failed",
                    }
                else:
                    responses = current_response.get("responses", []) if isinstance(current_response, dict) else []
                    result_table = responses[0].get("result_table", []) if responses and isinstance(responses[0], dict) else []
                    current_fields = result_table[0] if result_table and isinstance(result_table[0], dict) else {}

                    updates = diff_ticket_fields(desired_fields, current_fields)
                    if updates:
                        log.info(f"  ↻ 检测到字段变化，准备更新 {len(updates)} 项")
                        for diff_key in sorted(updates.keys())[:5]:
                            desired_norm = normalize_value_for_compare(diff_key, desired_fields.get(diff_key, ""))
                            current_norm = normalize_value_for_compare(diff_key, current_fields.get(diff_key, ""))
                            log.info(f"    - 变化字段 {diff_key}: current={current_norm!r} -> desired={desired_norm!r}")
                        updated, update_response, update_error = update_ticket_fields(
                            api_url=runtime_config["api_url"],
                            headers=headers,
                            timeout=args.timeout,
                            insecure=args.insecure,
                            tenant=tenant,
                            subject=subject,
                            ticket_id=row_ticket_id,
                            update_fields=updates,
                        )
                        if not updated:
                            result = {
                                "success": False,
                                "ticket_id": row_ticket_id,
                                "error": f"更新失败: {update_error}",
                                "payload": updates,
                                "action": "update_failed",
                            }
                        else:
                            update_esservice_error = extract_esservice_error(update_response)
                            if update_esservice_error:
                                result = {
                                    "success": False,
                                    "ticket_id": row_ticket_id,
                                    "error": f"更新失败: {update_esservice_error}",
                                    "payload": updates,
                                    "response": update_response,
                                    "action": "update_failed",
                                }
                            else:
                                log.info(f"  ✅ 更新成功  ticket_id={row_ticket_id}")
                                result = {
                                    "success": True,
                                    "ticket_id": row_ticket_id,
                                    "error": "",
                                    "payload": updates,
                                    "response": update_response,
                                    "action": "updated",
                                }
                    else:
                        log.info("  ℹ️ 无字段变化，跳过更新")
                        result = {
                            "success": True,
                            "ticket_id": row_ticket_id,
                            "error": "",
                            "payload": desired_fields,
                            "response": current_response,
                            "action": "no_change",
                        }
        else:
            result = create_ticket(
                row=row,
                api_url=runtime_config["api_url"],
                headers=headers,
                field_mapping=runtime_config["field_mapping"],
                static_fields=runtime_config["static_fields"],
                envelope=runtime_config["envelope"],
                response_path=runtime_config["response_path"],
                timeout=args.timeout,
                insecure=args.insecure,
                dry_run=args.dry_run,
            )
        result["row"] = index
        result["title"] = get_row_value(row, "Title")
        results.append(result)
        if not args.dry_run:
            time.sleep(args.pause_seconds)

    success_count = sum(1 for item in results if item["success"])
    fail_count = len(results) - success_count
    no_change_count = sum(1 for item in results if item.get("action") == "no_change")
    update_success_count = sum(1 for item in results if item.get("action") == "updated")
    update_failed_count = sum(1 for item in results if item.get("action") == "update_failed")
    created_count = sum(1 for item in results if item.get("action") == "created")

    log.info("=" * 60)
    log.info(f"完成: 成功={success_count}  失败={fail_count}  共={len(results)}")
    log.info(
        "分类统计: "
        f"无更新={no_change_count}  "
        f"更新成功={update_success_count}  "
        f"更新失败={update_failed_count}  "
        f"新建={created_count}"
    )
    save_results(args.excel, results)

    if not args.dry_run and success_count > 0:
        try:
            write_ticket_ids_to_excel(
                excel_path=args.excel,
                sheet_name=args.sheet,
                results=results,
            )
        except Exception as e:
            log.warning(f"⚠️ 回写Excel ticket_id列失败: {e}")

    if success_count > 0:
        log.info("=" * 60)
        log.info("✅ 创建成功的 Ticket ID:")
        for item in results:
            if item["success"]:
                ticket_id = item.get("ticket_id", "unknown")
                title = item.get("title", "无标题")
                log.info(f"  行 {item['row']}: {title} → ticket_id={ticket_id}")
        log.info("=" * 60)

    if fail_count > 0:
        for item in results:
            if not item["success"]:
                log.warning(f"行 {item['row']}: {item['title']} → {item['error']}")
        sys.exit(1)


if __name__ == "__main__":
    main()