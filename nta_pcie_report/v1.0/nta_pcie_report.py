# last updated: 260514_1200

import sys
import argparse
import logging
import re
import shutil
import datetime
import warnings
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import NamedStyle, Border, Side, PatternFill, Font, Alignment
from openpyxl.cell.cell import Cell, MergedCell


STLYE_DCT = {}
STLYE_DCT['text_style'] = NamedStyle(name='text_style', number_format='@')
STLYE_DCT['thin_border'] = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
STLYE_DCT['fill_pass'] = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
STLYE_DCT['font_pass'] = Font(color='00B050')
STLYE_DCT['fill_fail'] = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
STLYE_DCT['font_fail'] = Font(color='FF0000')
STLYE_DCT['fill_na'] = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
STLYE_DCT['center_align'] = Alignment(horizontal='center', vertical='top')
STLYE_DCT['left_align'] = Alignment(horizontal='left', vertical='top')
STLYE_DCT['fill_brown'] = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
STLYE_DCT['font_brown'] = Font(color='806000')
STLYE_DCT['fill_skyblue'] = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
STLYE_DCT['font_skyblue'] = Font(color='0070C0')
STLYE_DCT['fill_purple'] = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
STLYE_DCT['font_purple'] = Font(color='833C0C')
STLYE_DCT['fill_yellow'] = PatternFill(start_color='FFFF66', end_color='FFFF66', fill_type='solid')
STLYE_DCT['font_yellow'] = Font(color='0070C0')


class LtssmStress:
    def __init__(self, summary_xlsx_lst: list[Path]) -> None:
        self.summary_xlsx_lst = summary_xlsx_lst
        self.log_pairs_lst: list[tuple[Path, Path | None]] = []
        self.ltssm_no_debug_df_dct: dict[Path, pd.DataFrame] = {}
        self.ltssm_debug_df_dct: dict[Path, pd.DataFrame] = {}
        self.ltssm_fn_ids_dct: dict[Path, tuple] = {}
        self.debug_pxp_x8_lst: list[str] = []

    def create_log_pairs(self):
        for summary_xlsx in self.summary_xlsx_lst:
            l1_off_log_name = re.sub(r'_\d{6}', '', summary_xlsx.name.split('ltloop_summary')[0])

            l1_on_runpml_log_lst = [f for f in summary_xlsx.parent.parent.rglob('*runPML1_itplog*') if not f.name.startswith('~') and 'deprecated' not in str(f)]
            result_lst: list[Path] = []
            for l1_on_runpml_log in l1_on_runpml_log_lst:
                check_name = re.sub(r'_\d{6}', '', l1_on_runpml_log.name.split('runPML1_itplog')[0])
                if check_name == l1_off_log_name:
                    result_lst.append(l1_on_runpml_log)

            if len(result_lst) == 0:
                warning_item_msg(summary_xlsx, 'There is no "runPML1_itplog" log matching this "ltloop_summary" log')
                self.log_pairs_lst.append((summary_xlsx, None))
            elif len(result_lst) > 1:
                error_item_msg(summary_xlsx, 'More than 1 runPML1_itplog log matching this ltloop_summary log', no_exit=True)
                logging.info(list(map(lambda x: x.name, result_lst)))
                sys.exit(1)
            else:
                self.log_pairs_lst.append((summary_xlsx, result_lst[0]))

    def generate_ltssm_df(self):
        for summary_xlsx, l1_on_runpml_log in self.log_pairs_lst:
            ori_df = pd.read_excel(summary_xlsx, index_col=False)
            l1_off_runpml_df = ori_df.drop(columns=['index', 'Start Time', 'Stop Time'])
            if l1_on_runpml_log:
                l1_on_runpml_row = self.extract_l1_on_runpml_info(l1_on_runpml_log)
                l1_on_runpml_df = pd.DataFrame(l1_on_runpml_row).T
                l1_on_runpml_df.columns = l1_off_runpml_df.columns
                ltssm_df = pd.concat([l1_off_runpml_df, l1_on_runpml_df], ignore_index=True)
            else:
                ltssm_df = l1_off_runpml_df
            if all('debug' not in part.lower() for part in summary_xlsx.resolve().parts):
                self.ltssm_no_debug_df_dct[summary_xlsx] = ltssm_df
            else:
                if not l1_on_runpml_log:
                    ltssm_df.loc[len(ltssm_df)] = None
                self.ltssm_debug_df_dct[summary_xlsx] = ltssm_df
        self.ltsmm_gen_pxp_port_from_fn()

    def extract_l1_on_runpml_info(self, l1_on_runpml_log: Path):
        log_content = l1_on_runpml_log.read_text(encoding='utf-8', errors='ignore')
        l1_on_runpml_result = self.generate_l1_on_runpml_result(log_content, l1_on_runpml_log)
        count = self.generate_l1_on_runpml_count(log_content, l1_on_runpml_log)
        duration = self.generate_l1_on_runpml_duration(log_content, l1_on_runpml_log)
        row = ['pml1', l1_on_runpml_result, duration, count] + ['NA' ] * 4
        return row

    @staticmethod
    def generate_l1_on_runpml_result(log_content: str, l1_on_runpml_log: Path):
        summary_idx = log_content.find('SUMMARY:')
        if summary_idx != -1:
            log_summary = log_content[summary_idx:].lstrip()
            error_pairs = re.findall(r'(\d+)% or (\d+) errors', log_summary)
            runpml_has_error = any(int(p) > 0 or int(e) != 0 for p, e in error_pairs)
        else:
            runpml_has_error = True
            error_item_msg(l1_on_runpml_log, 'No Summary block in this log')
        return 'PASS' if not runpml_has_error else 'FAIL'

    @staticmethod
    def generate_l1_on_runpml_count(log_content: str, l1_on_runpml_log: Path):
        count_idx = log_content.rfind("---------Starting runPML1 iteration ")
        if count_idx != -1:
            line_end = log_content.find('\n', count_idx)
            count_line = log_content[count_idx:line_end if line_end != -1 else None]
            try:
                count = int(count_line.strip('-').split()[-1])
            except ValueError:
                logging.warning('Failed to get count')
                count = 'NA'
        else:
            count = 0
            error_item_msg(l1_on_runpml_log, 'No Starting runPML1 iteration in this log')
        return count

    @staticmethod
    def generate_l1_on_runpml_duration(log_content: str, l1_on_runpml_log: Path):
        test_time_idx = log_content.rfind("Total Test Time")
        if test_time_idx != -1:
            duration_st_idx = log_content.find('\n',test_time_idx) + 1
            duration_str = log_content[duration_st_idx:].splitlines()[0]
            duration_part = duration_str.split()
            d = float(duration_part[0])
            h, m, s = map(float, duration_part[2].split(':'))
            duration = d * 24 * 60 + h * 60 + m + s / 60
        else:
            duration = 0.0
            error_item_msg(l1_on_runpml_log, 'No Total Test Time in this log')
        return duration

    def ltsmm_gen_pxp_port_from_fn(self):
        for summary_xlsx in self.summary_xlsx_lst:
            gen = self.extract_info_from_fn(r'Gen\d+', summary_xlsx)
            pxp = self.extract_info_from_fn(r'pxp\d+', summary_xlsx)
            port = self.extract_info_from_fn(r'port\d+', summary_xlsx).replace('port', '')
            self.ltssm_fn_ids_dct[summary_xlsx] = (gen, pxp, int(port))

    @staticmethod
    def extract_info_from_fn(pattern: str, file_path: Path):
        re_search_resule = re.search(f'{pattern}', file_path.name)
        if re_search_resule:
            return re_search_resule.group()
        err_msg = re.sub(r'\+\\d', 'X', pattern)
        error_item_msg(file_path, f'The "ltloop_summary" file name does not contain "{err_msg}"')
        return ''


def write_df_to_xlsx(df: pd.DataFrame, ws: Worksheet, st_row, st_col, pxp_sp=None):
    try:
        for row_idx, row in enumerate(df.values, start=st_row):
            for col_idx, value in enumerate(row, start=st_col):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, float):
                    cell.number_format = '0.00'
                else:
                    if not cell.style:
                        cell.style = STLYE_DCT['text_style']
                cell.border = STLYE_DCT['thin_border']
                cell.alignment = STLYE_DCT['left_align']
                xlsx_sp_value_style(cell, value, col_idx, pxp_sp)
                if col_idx == 3:
                    cell.alignment = STLYE_DCT['center_align']
                # xlsx_sp_col_idx_style(cell, value, col_idx)
    except AttributeError as error_obj:
        if "'MergedCell' object attribute 'value' is read-only" in error_obj.args:
            logging.error('Please check if there are any additional merged cells in the report file.')
        sys.exit(1)


def xlsx_sp_value_style(cell: Cell | MergedCell, value, col_idx, pxp_sp=None):
    match value:
        case 'PASS':
            cell.fill = STLYE_DCT['fill_pass']
            cell.font = STLYE_DCT['font_pass']
        case 'FAIL':
            cell.fill = STLYE_DCT['fill_fail']
            cell.font = STLYE_DCT['font_fail']
        case 'NA':
            cell.fill = STLYE_DCT['fill_na']
        case 'Gen5' if col_idx == 1:
            cell.fill = STLYE_DCT['fill_skyblue']
            cell.font = STLYE_DCT['font_skyblue']
        case 'Gen6' if col_idx == 1:
            cell.fill = STLYE_DCT['fill_brown']
            cell.font = STLYE_DCT['font_brown']
        case 0 if col_idx == 3:
            cell.fill = STLYE_DCT['fill_brown']
            cell.font = STLYE_DCT['font_brown']
        case 4 if col_idx == 4:
            cell.fill = STLYE_DCT['fill_skyblue']
            cell.font = STLYE_DCT['font_skyblue']
        case _ if col_idx == 2 and pxp_sp == 'x8' and re.fullmatch(r'pxp\d+', value):
            cell.fill = STLYE_DCT['fill_purple']
            cell.font = STLYE_DCT['font_purple']
        case _ if col_idx == 2 and pxp_sp == 'x16' and re.fullmatch(r'pxp\d+', value):
            cell.fill = STLYE_DCT['fill_yellow']
            cell.font = STLYE_DCT['font_yellow']


def write_ltsmm_summary(ltssm_stress_obj: LtssmStress, report_xlsx: Path):
    warnings.filterwarnings('ignore', category=UserWarning)
    wb = load_workbook(report_xlsx)
    sheets = wb.sheetnames
    if 'LTSSM_Stress' not in sheets:
        error_item_msg('LTSSM_Stress', 'No such sheet', close_wb=wb)
    ws = wb['LTSSM_Stress']
    ori_df = pd.DataFrame(ws.values)

    ltssm_stress.generate_ltssm_df()

    for xlsx_path, df in ltssm_stress_obj.ltssm_no_debug_df_dct.items():
        st_row = extract_ltsmm_df_position(ori_df, ltssm_stress_obj.ltssm_fn_ids_dct[xlsx_path])
        st_col_test = (ori_df.iloc[0] == 'Test').to_numpy().nonzero()[0][0] + 1
        st_col_duration = (ori_df.iloc[0].astype(str).str.contains('Duration')).to_numpy().nonzero()[0][0] + 1
        write_df_to_xlsx(df.iloc[:, :2], ws, st_row, st_col_test)
        write_df_to_xlsx(df.iloc[:, 2:], ws, st_row, st_col_duration)
    if ltssm_stress_obj.ltssm_debug_df_dct:
        write_ltsmm_debug(ori_df, ltssm_stress_obj, ws)
    wb.save(report_xlsx)
    wb.close()


def write_ltsmm_debug(ori_df: pd.DataFrame, ltssm_stress_obj: LtssmStress, ws: Worksheet):
    debug_row = get_debug_row(ori_df)
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= debug_row:
            ws.unmerge_cells(str(rng))
    for xlsx_path, df in ltssm_stress_obj.ltssm_debug_df_dct.items():
        pxp = 'x8' if ltssm_stress_obj.ltssm_fn_ids_dct[xlsx_path][2] == 4 else 'x16'
        df_height = len(df)
        new_df_prefix_lst = [list(ltssm_stress_obj.ltssm_fn_ids_dct[xlsx_path]) + ['', 'Link L1 Disable'] for _ in range(df_height)]
        new_df_prefix_lst[-1][-1] = 'Link L1 Enable'
        new_df_prefix = pd.DataFrame(new_df_prefix_lst, columns=['Gen#', 'Pxpstr', 'Portno', 'WO#', ''])
        new_df_suffix_lst = [['', '', ''] for _ in range(df_height)]
        new_df_suffix_lst[0][-1] = 'debug'
        new_df_suffix = pd.DataFrame(new_df_suffix_lst, columns=['HSD', 'HSD Status', ''])
        st_col_duration = (ori_df.iloc[0].astype(str).str.contains('Duration')).to_numpy().nonzero()[0][0] + 1
        write_df_to_xlsx(pd.concat([new_df_prefix, df.iloc[:, :2]], axis=1), ws, debug_row, 1, pxp_sp=pxp)
        write_df_to_xlsx(pd.concat([df.iloc[:, 2:], new_df_suffix], axis=1), ws, debug_row, st_col_duration)
        ws.merge_cells(start_row=debug_row, start_column=3, end_row=debug_row + df_height - 1, end_column=3)
        ws.merge_cells(start_row=debug_row, start_column=4, end_row=debug_row + df_height - 1, end_column=4)
        ws.merge_cells(start_row=debug_row, start_column=5, end_row=debug_row + df_height - 2, end_column=5)
        debug_row += df_height


def get_debug_row(ori_df: pd.DataFrame):
    mask = ori_df.map(lambda x: 'debug' in str(x))
    match_rows = mask.any(axis=1)
    debug_row = int(match_rows.idxmax() if match_rows.any() else ori_df.index[-1] + 1) + 1
    return debug_row


def extract_ltsmm_df_position(df: pd.DataFrame, ids: tuple):
    # ids = (gen, pxp, port)
    gen, pxp, port = ids
    gen_1st_line = df[df[0] == gen].index[0]
    pxp_1st_line = df[(df.index >= gen_1st_line) & (df[1] == pxp)].index[0]
    st_row = df[(df.index >= pxp_1st_line)& (df[2] == port)].index[0] + 1
    return st_row


def error_item_msg(item, msg, close_wb: Workbook | None = None, no_exit=False):
    logging.error('%s:', msg)
    logging.error('%s', item)
    if close_wb:
        close_wb.close()
    if no_exit:
        return
    sys.exit(1)


def warning_item_msg(item, msg):
    logging.warning('%s:', msg)
    logging.warning('%s', item)


def get_last_update():
    with open(__file__, 'r', encoding='utf-8') as f:
        for line in f:
            if '# last update' in line:
                print(line.strip('#').strip())
                sys.exit(0)
    print('Unknow version.')
    sys.exit(0)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description='Updated the NTA PCIE report.')
    parser.add_argument('ori_xlsx', nargs='?', help='Specify the report file.')
    parser.add_argument('log_folder', nargs='?', help='Specify the logs folder.')
    parser.add_argument('-v', '--version', action='store_true', help='Display the version.')
    args = parser.parse_args(argv)

    if args.version:
        get_last_update()
    if args.ori_xlsx is None or args.log_folder is None:
        parser.print_help()
        sys.exit(0)
    return args, parser


logging.basicConfig(level=logging.DEBUG, format='[%(asctime)s] [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logging.getLogger('PIL').setLevel(logging.WARNING)
parsed_args, args_parser = parse_args()

log_folder = Path(parsed_args.log_folder)
ori_xlsx = Path(parsed_args.ori_xlsx).resolve()

if not log_folder.exists():
    error_item_msg(log_folder, 'No such directory')
elif not log_folder.is_dir():
    logging.error('"log_folder" must be the folder.')
    args_parser.print_help()
    sys.exit(1)
if ori_xlsx.exists():
    if ori_xlsx.suffix != '.xlsx':
        logging.error('"ori_xlsx" must be the xlsx file.')
        args_parser.print_help()
        sys.exit(1)
    try:
        bak_xlsx = ori_xlsx.parent / f'{ori_xlsx.stem}_bak_{datetime.datetime.now().strftime("%y%m%d_%H%M%S")}.xlsx'
        shutil.copy(ori_xlsx, bak_xlsx)
        logging.info('Report file backuped:')
        logging.info('%s', bak_xlsx)
    except PermissionError:
        logging.error('Please confirm that the file is not opened by any other program')
        sys.exit(1)
else:
    error_item_msg(ori_xlsx, 'No such file')
ltsmm_summary_xlsx_lst = [f for f in log_folder.rglob('*ltloop_summary*') if not f.name.startswith('~') and 'deprecated' not in str(f)]
ltssm_stress = LtssmStress(ltsmm_summary_xlsx_lst)
ltssm_stress.create_log_pairs()
write_ltsmm_summary(ltssm_stress, ori_xlsx)
